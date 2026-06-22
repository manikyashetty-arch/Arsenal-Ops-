"""
Integration tests for the roadmap router (POST/GET /api/roadmap/*).

Tests cover XLSX upload, file validation, sprint extraction, and parsing.
Mocks AI LLM calls; uses openpyxl to generate minimal valid XLSX files in-memory.

Fixtures from conftest.py: db, test_client, make_token, admin_user, pm_user, dev_user
"""

import io
from datetime import datetime, timedelta
from unittest import mock

import pytest
from openpyxl import Workbook

from models.developer import Developer
from tests.conftest import seed_project

# ============= Test Helpers =============


def create_minimal_xlsx() -> bytes:
    """Create a minimal valid XLSX file with roadmap structure.

    Returns bytes suitable for upload (via BytesIO).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "roadmap"

    # Header row
    headers = [
        "Type",
        "Name",
        "Description",
        "Milestone",
        "Epic",
        "Priority",
        "Effort (hrs)",
        "Assignee",
    ]
    ws.append(headers)

    # Add a milestone
    ws.append(["MILESTONE", "Phase 1", None, None, None, None, None, None])

    # Add week dates (columns I, J, K, L, M)
    # These are Monday dates for weeks 1-5
    base_date = datetime(2026, 6, 1)  # A Monday
    for i in range(5):
        week_date = base_date + timedelta(weeks=i)
        ws.cell(row=2, column=9 + i, value=week_date)

    # Add an epic
    ws.append(
        [
            "EPIC",
            "Feature A",
            "Description of Feature A",
            "Phase 1",
            "Feature A",
            "High",
            None,
            None,
        ]
    )

    # Add tasks with hours
    ws.append(
        [
            "TASK",
            "Task 1",
            "First task",
            "Phase 1",
            "Feature A",
            "High",
            16.0,
            "Developer 1",
            8,
            8,
            0,
            0,
            0,
        ]
    )
    ws.append(
        [
            "TASK",
            "Task 2",
            "Second task",
            "Phase 1",
            "Feature A",
            "Medium",
            20.0,
            "Developer 2",
            0,
            0,
            10,
            10,
            0,
        ]
    )
    ws.append(
        [
            "TASK",
            "Task 3",
            "Third task",
            "Phase 1",
            "Feature A",
            "Low",
            12.0,
            "Developer 1",
            0,
            0,
            6,
            6,
            0,
        ]
    )

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_minimal_xlsx_two_sprints() -> bytes:
    """Create XLSX with 2 distinct sprints (4 weeks total at 2 weeks per sprint).

    This is used to test sprint extraction and boundary validation.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "roadmap"

    # Header (9 columns for type, name, desc, milestone, epic, priority, effort, assignee, + 4 week columns)
    headers = [
        "Type",
        "Name",
        "Description",
        "Milestone",
        "Epic",
        "Priority",
        "Effort (hrs)",
        "Assignee",
        "W1",
        "W2",
        "W3",
        "W4",
    ]
    ws.append(headers)

    # Milestone row with all 4 weeks
    ws.append(["MILESTONE", "Phase 1", None, None, None, None, None, None])
    base_date = datetime(2026, 6, 1)  # Monday
    for i in range(4):
        week_date = base_date + timedelta(weeks=i)
        ws.cell(row=2, column=9 + i, value=week_date)

    # Epic for sprint 1
    ws.append(
        [
            "EPIC",
            "Feature Sprint1",
            "Epic for sprint 1",
            "Phase 1",
            "Feature Sprint1",
            "High",
            None,
            None,
        ]
    )

    # Tasks for sprint 1 (active in weeks 1-2)
    ws.append(
        [
            "TASK",
            "Task Sprint1-A",
            "Task A",
            "Phase 1",
            "Feature Sprint1",
            "High",
            16.0,
            "Alice",
            8,
            8,
            0,
            0,
        ]
    )
    ws.append(
        [
            "TASK",
            "Task Sprint1-B",
            "Task B",
            "Phase 1",
            "Feature Sprint1",
            "Medium",
            8.0,
            "Bob",
            4,
            4,
            0,
            0,
        ]
    )

    # Epic for sprint 2
    ws.append(
        [
            "EPIC",
            "Feature Sprint2",
            "Epic for sprint 2",
            "Phase 1",
            "Feature Sprint2",
            "Low",
            None,
            None,
        ]
    )

    # Tasks for sprint 2 (active in weeks 3-4)
    ws.append(
        [
            "TASK",
            "Task Sprint2-A",
            "Task A Sprint 2",
            "Phase 1",
            "Feature Sprint2",
            "Medium",
            12.0,
            "Charlie",
            0,
            0,
            6,
            6,
        ]
    )
    ws.append(
        [
            "TASK",
            "Task Sprint2-B",
            "Task B Sprint 2",
            "Phase 1",
            "Feature Sprint2",
            "Low",
            10.0,
            "Alice",
            0,
            0,
            5,
            5,
        ]
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============= Authentication Tests =============


class TestRoadmapAuth:
    """Tests for auth requirements on roadmap endpoints."""

    def test_upload_requires_auth(self, db, test_client):
        """Verify POST /parse-file without token → 401.

        Omit Authorization header; assert 401.
        """
        project = seed_project(db, "Test Project", num_developers=1)

        # Upload without token
        xlsx_bytes = create_minimal_xlsx()
        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 401


# ============= File Validation Tests =============


class TestRoadmapFileValidation:
    """Tests for file type and format validation."""

    def _setup_project_with_user(self, db, admin_user):
        """Helper: create a project and ensure user is associated."""
        _user, token = admin_user
        project = seed_project(db, "Upload Test", num_developers=1)
        return project, token

    def test_upload_invalid_file_type_returns_400(self, db, test_client, admin_user):
        """Verify POST with non-XLSX file → 400.

        Send PDF bytes (fake) or text file; assert 400 + "Excel file" message.
        """
        project, token = self._setup_project_with_user(db, admin_user)

        # Send fake PDF
        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={"file": ("test.pdf", b"fake pdf content", "application/pdf")},
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 400
        assert "Excel file" in response.json()["detail"]

    def test_upload_empty_file_returns_400(self, db, test_client, admin_user):
        """Verify POST with empty file body → 400.

        Send empty bytes; assert 4xx.
        """
        project, token = self._setup_project_with_user(db, admin_user)

        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "empty.xlsx",
                    b"",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert 400 <= response.status_code < 500

    def test_upload_invalid_sprint_weeks_returns_400(self, db, test_client, admin_user):
        """Verify POST with sprint_weeks outside [1, 6] → 400.

        Test with sprint_weeks=0, sprint_weeks=7; assert 400.
        """
        project, token = self._setup_project_with_user(db, admin_user)
        xlsx_bytes = create_minimal_xlsx()

        # Too low
        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "0"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 400
        assert "Sprint weeks" in response.json()["detail"]

        # Too high
        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "7"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 400
        assert "Sprint weeks" in response.json()["detail"]

    def test_upload_nonexistent_project_returns_404(self, db, test_client, admin_user):
        """Verify POST with nonexistent project_id → 404.

        Use project_id=999999; assert 404 + "Project not found".
        """
        _, token = admin_user
        xlsx_bytes = create_minimal_xlsx()

        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": "999999", "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 404
        assert "Project not found" in response.json()["detail"]


# ============= Happy Path Tests =============


class TestRoadmapParsing:
    """Tests for successful roadmap parsing."""

    def _setup_project_with_user(self, db, admin_user):
        """Helper: create project and return token."""
        _user, token = admin_user
        project = seed_project(db, "Parse Test", num_developers=2)
        return project, token

    @mock.patch("routers.roadmap.get_roadmap_ai_parser")
    def test_upload_xlsx_returns_success_payload(self, mock_ai_parser, db, test_client, admin_user):
        """Verify POST with valid XLSX → 200 + RoadmapParseResponse.

        Uploads minimal XLSX, asserts 200, and checks response has:
        - status: "success"
        - summary: with total_epics, total_tasks, total_assignees, timeline, conflicts, warnings
        - parsed_data: dict
        """
        project, token = self._setup_project_with_user(db, admin_user)
        xlsx_bytes = create_minimal_xlsx()

        # Mock AI parser to not be called (standard parser should succeed)
        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "success"

        summary = data["summary"]
        assert "total_epics" in summary
        assert "total_tasks" in summary
        assert "total_assignees" in summary
        assert "timeline" in summary
        assert "conflicts" in summary
        assert "warnings" in summary
        assert "schedule" in summary

        parsed = data["parsed_data"]
        assert "tickets" in parsed
        assert "meta" in parsed

    @mock.patch("routers.roadmap.get_roadmap_ai_parser")
    def test_upload_counts_tickets_and_epics(self, mock_ai_parser, db, test_client, admin_user):
        """Verify parsed summary counts match input.

        Creates XLSX with 3 tasks, 1 epic; assert summary reflects this.
        """
        project, token = self._setup_project_with_user(db, admin_user)
        xlsx_bytes = create_minimal_xlsx()

        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        summary = response.json()["summary"]

        # Our minimal XLSX has 3 tasks and 1 epic
        assert summary["total_tasks"] == 3
        assert summary["total_epics"] == 1
        # 2 assignees (Developer 1, Developer 2)
        assert summary["total_assignees"] == 2


# ============= Sprint Extraction Tests =============


class TestSprintExtraction:
    """Tests for sprint calculation and boundaries."""

    def _setup_project_with_user(self, db, admin_user):
        """Helper: create project and return token."""
        _user, token = admin_user
        project = seed_project(db, "Sprint Test", num_developers=3)
        return project, token

    @mock.patch("routers.roadmap.get_roadmap_ai_parser")
    def test_extracted_sprints_have_valid_boundaries(
        self, mock_ai_parser, db, test_client, admin_user
    ):
        """Verify sprints have start < end, non-overlapping, ordered by start.

        Creates XLSX with 2 sprints (4 weeks, 2 weeks per sprint).
        Asserts:
        - Each sprint's start_date < end_date
        - Sprints are ordered by start_date
        - No overlap between consecutive sprints
        """
        project, token = self._setup_project_with_user(db, admin_user)
        xlsx_bytes = create_minimal_xlsx_two_sprints()

        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        parsed = response.json()["parsed_data"]
        sprints = parsed.get("sprints", [])

        # Should have 2 sprints
        assert len(sprints) >= 2, f"Expected at least 2 sprints, got {len(sprints)}"

        # Verify boundaries and ordering
        prev_end = None
        for i, sprint in enumerate(sprints):
            start = sprint["start_week"]
            end = sprint["end_week"]

            # Verify start < end
            start_dt = datetime.fromisoformat(start)
            end_dt = datetime.fromisoformat(end)
            assert start_dt < end_dt, f"Sprint {i}: start {start} >= end {end}"

            # Verify ordering
            if prev_end:
                prev_end_dt = datetime.fromisoformat(prev_end)
                assert start_dt >= prev_end_dt, (
                    f"Sprint {i}: not ordered by start; prev_end={prev_end}, start={start}"
                )

            prev_end = end

    @mock.patch("routers.roadmap.get_roadmap_ai_parser")
    def test_sprints_contain_sprint_number(self, mock_ai_parser, db, test_client, admin_user):
        """Verify each sprint has a number field and tasks list."""
        project, token = self._setup_project_with_user(db, admin_user)
        xlsx_bytes = create_minimal_xlsx_two_sprints()

        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        sprints = response.json()["parsed_data"]["sprints"]

        for sprint in sprints:
            assert "number" in sprint
            assert isinstance(sprint["number"], int)
            assert "tasks" in sprint
            assert isinstance(sprint["tasks"], list)


# ============= Commit Endpoint Tests =============


class TestRoadmapCommit:
    """Tests for POST /commit endpoint (creating work items from parsed roadmap)."""

    def _setup_project_with_dev(self, db, admin_user):
        """Helper: create project with developers, return (project, token, admin_user)."""
        user, token = admin_user

        # Ensure user is a developer
        dev = db.query(Developer).filter(Developer.email == user.email).first()
        if not dev:
            dev = Developer(
                name=user.name,
                email=user.email,
                github_username="admin-dev",
            )
            db.add(dev)
            db.commit()

        project = seed_project(db, "Commit Test", num_developers=2)
        return project, token

    def test_commit_requires_auth(self, db, test_client):
        """Verify POST /commit without token → 401."""
        response = test_client.post(
            "/api/roadmap/commit",
            json={"project_id": 1, "parsed_data": {}},
        )

        assert response.status_code == 401

    def test_commit_nonexistent_project_returns_404(self, db, test_client, admin_user):
        """Verify POST /commit with invalid project_id → 404."""
        _, token = admin_user

        response = test_client.post(
            "/api/roadmap/commit",
            json={"project_id": 999999, "parsed_data": {}},
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 404
        assert "Project not found" in response.json()["detail"]

    @mock.patch("routers.roadmap.get_roadmap_ai_parser")
    def test_commit_creates_work_items_from_parsed_data(
        self, mock_ai_parser, db, test_client, admin_user
    ):
        """Verify POST /commit creates epics and tasks from parsed_data.

        First parse a roadmap, then commit it; verify work items are created in db.
        """
        project, token = self._setup_project_with_dev(db, admin_user)
        xlsx_bytes = create_minimal_xlsx()

        # Parse the file first
        parse_response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert parse_response.status_code == 200
        parsed_data = parse_response.json()["parsed_data"]

        # Now commit
        commit_response = test_client.post(
            "/api/roadmap/commit",
            json={"project_id": project.id, "parsed_data": parsed_data},
            headers={"Authorization": f"Bearer {token}"},
        )

        assert commit_response.status_code == 200
        response_data = commit_response.json()
        assert response_data["status"] == "success"
        assert response_data["tickets_created"] == 3
        assert response_data["epics_created"] == 1


# ============= AI Parser Fallback Tests =============


class TestAIParserFallback:
    """Tests for fallback to AI parser when standard parser fails."""

    def _setup_project_with_user(self, db, admin_user):
        """Helper: create project and return token."""
        _user, token = admin_user
        project = seed_project(db, "AI Fallback Test", num_developers=2)
        return project, token

    def test_upload_calls_ai_parser_on_standard_failure(self, db, test_client, admin_user):
        """Verify AI parser is called when standard parser fails.

        Send a malformed XLSX that standard parser will reject.
        Mock the AI parser to return a valid result.
        Assert AI parser was called.
        """
        project, token = self._setup_project_with_user(db, admin_user)

        # Create a minimal but invalid XLSX (just empty sheet)
        wb = Workbook()
        ws = wb.active
        ws.title = "roadmap"
        # No data at all

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        invalid_xlsx = output.getvalue()

        # Mock the AI parser
        mock_parsed = {
            "meta": {
                "file": "test.xlsx",
                "parsed_at": datetime.now().isoformat(),
                "week_range": {"start": "2026-06-01", "end": "2026-06-08"},
                "total_weeks": 1,
                "total_tasks": 0,
                "total_assignees": 0,
                "total_sprints": 0,
            },
            "tickets": [],
            "schedule": {},
            "conflicts": [],
            "parallel_tasks": [],
            "availability": {},
            "warnings": [{"row": 0, "task": "none", "issue": "unassigned", "detail": "Empty file"}],
            "sprints": [],
            "unscheduled_tasks": [],
        }

        with mock.patch("routers.roadmap.get_roadmap_ai_parser") as mock_get_parser:
            mock_parser_instance = mock.AsyncMock()
            mock_parser_instance.parse_excel_with_ai = mock.AsyncMock(return_value=mock_parsed)
            mock_get_parser.return_value = mock_parser_instance

            response = test_client.post(
                "/api/roadmap/parse-file",
                data={"project_id": str(project.id), "sprint_weeks": "2"},
                files={
                    "file": (
                        "test.xlsx",
                        invalid_xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                headers={"Authorization": f"Bearer {token}"},
            )

            # Should succeed via AI parser
            assert response.status_code == 200
            assert response.json()["parsed_data"]["meta"]["parser_used"] == "ai"


# ============= Conflict and Warning Detection Tests =============


class TestConflictDetection:
    """Tests for identifying scheduling conflicts and warnings."""

    def _setup_project_with_user(self, db, admin_user):
        """Helper: create project and return token."""
        _user, token = admin_user
        project = seed_project(db, "Conflict Test", num_developers=2)
        return project, token

    @mock.patch("routers.roadmap.get_roadmap_ai_parser")
    def test_parse_detects_unassigned_tasks(self, mock_ai_parser, db, test_client, admin_user):
        """Verify parser flags unassigned tasks in warnings.

        Create XLSX with a task but no assignee; assert warnings contain 'unassigned'.
        """
        project, token = self._setup_project_with_user(db, admin_user)

        # Create XLSX with unassigned task
        wb = Workbook()
        ws = wb.active
        ws.title = "roadmap"

        headers = [
            "Type",
            "Name",
            "Description",
            "Milestone",
            "Epic",
            "Priority",
            "Effort (hrs)",
            "Assignee",
        ]
        ws.append(headers)
        ws.append(["MILESTONE", "Phase 1", None, None, None, None, None, None])

        base_date = datetime(2026, 6, 1)
        for i in range(2):
            week_date = base_date + timedelta(weeks=i)
            ws.cell(row=2, column=9 + i, value=week_date)

        ws.append(["EPIC", "Feature A", "Desc", "Phase 1", "Feature A", "High", None, None])
        # Task with no assignee (empty or None)
        ws.append(
            [
                "TASK",
                "Unassigned Task",
                "No assignee",
                "Phase 1",
                "Feature A",
                "High",
                8.0,
                None,
                8,
                0,
            ]
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        xlsx_bytes = output.getvalue()

        response = test_client.post(
            "/api/roadmap/parse-file",
            data={"project_id": str(project.id), "sprint_weeks": "2"},
            files={
                "file": (
                    "test.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        warnings = response.json()["parsed_data"]["warnings"]

        # Should have at least one unassigned warning
        unassigned_warnings = [w for w in warnings if w.get("issue") == "unassigned"]
        assert len(unassigned_warnings) > 0


# ============= Xfail and Skip Tests =============


class TestKnownIssues:
    """Tests for known issues marked as xfail or skipped."""

    @pytest.mark.skip(
        reason="openpyxl formula evaluation not fully supported in test environment; safe to skip"
    )
    def test_upload_with_formula_cells(self, db, test_client, admin_user):
        """Verify upload handles Excel formulas gracefully.

        Skipped: formula evaluation is environment-dependent.
        """
        pass
