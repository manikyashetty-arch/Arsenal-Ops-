"""End-to-end tests for the Roadmap Parser feature.

Covers TWO surfaces:

  1. `parser.parse(filepath, sprint_weeks)` — the structured Excel parser
     that turns a roadmap workbook into the canonical `parsed_data` dict
     (tickets, schedule, sprints, conflicts, warnings, etc.). Tested by
     building real xlsx files on disk via openpyxl and feeding them through
     the parser.

  2. `routers.roadmap.commit_roadmap_tickets` — the endpoint handler that
     materialises that dict into the DB (epics + tasks + sprints, hour
     rollups, sprint assignment, activity log). Called directly with an
     in-memory SQLite session — the same pattern as
     `test_epic_hour_rollup.py`. We bypass FastAPI's `Depends()` machinery
     by passing `db` and `current_user` explicitly; this keeps the test
     focused on the function's own logic and skips the auth layer (which
     has its own tests).

Test plan (mapped to the user's spec):

  - File format / structure validation
  - Column-header detection (canonical + variations + missing)
  - Ticket / epic / milestone extraction
  - Sprint computation (windowing, Friday-end-week, multi-sprint)
  - Hour bookkeeping (per-week, per-assignee, conflicts, parallel tasks)
  - Warnings (unassigned, effort mismatch, no-weeks-planned, tolerance)
  - Commit: hierarchy (do child items belong to the right parent?)
  - Commit: hour rollup (do child hours sum to parent hours?)
  - Commit: assignee resolution + fallback
  - Commit: sprint assignment + due dates
  - Commit: activity log
  - Commit: 404 / 400 error paths
  - Edge cases (empty book, bad effort cell, mixed case row types, etc.)
"""

import datetime as dt
import os
import sys

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from database import Base
from models import (  # noqa: F401
    activity_log,
    architecture,
    developer,
    market_insight,
    persona,
    personal_task,
    project,
    project_file,
    project_goal,
    project_milestone,
    sprint,
    task,
    task_dependency,
    time_entry,
    user,
    user_story,
    work_item,
)
from models.activity_log import ActivityLog
from models.developer import Developer
from models.project import Project
from models.sprint import Sprint
from models.user import User
from models.work_item import WorkItem
from parser import parse as parse_roadmap
from routers.roadmap import RoadmapCommitRequest, commit_roadmap_tickets

# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────

WEEK_DATES = [
    dt.date(2026, 1, 5),  # Mon W1
    dt.date(2026, 1, 12),  # Mon W2
    dt.date(2026, 1, 19),  # Mon W3
    dt.date(2026, 1, 26),  # Mon W4
]


def _w(idx):
    """Return ISO-string for week idx (0-based) — convenience for assertions."""
    return WEEK_DATES[idx].isoformat()


@pytest.fixture
def db():
    """Fresh in-memory SQLite session per test."""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    session = Session()
    yield session
    session.close()


@pytest.fixture
def workbook_factory(tmp_path):
    """Returns a function that builds an xlsx on disk and yields its path.

    The factory signature is ``make(rows, sheet_name="Roadmap")`` where
    ``rows`` is a list of lists (first row = headers). Any cell whose value
    is a `datetime.date` is written as such so the parser's date detection
    fires.
    """

    counter = {"n": 0}

    def _make(rows, sheet_name="Roadmap"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        for row in rows:
            ws.append(row)
        counter["n"] += 1
        path = tmp_path / f"book_{counter['n']}.xlsx"
        wb.save(path)
        return str(path)

    return _make


def _canonical_header():
    """The 8-column left-side schema + 4 week date headers — closely matches
    what `services.roadmap_generator.build_xlsx` produces."""
    return [
        "Type",
        "Name",
        "Description",
        "Milestone",
        "Epic",
        "Priority",
        "Effort (hrs)",
        "Assignee",
        WEEK_DATES[0],
        WEEK_DATES[1],
        WEEK_DATES[2],
        WEEK_DATES[3],
    ]


def _basic_rows():
    """Two epics, four tasks, one milestone, four weeks — used as the happy
    baseline by multiple tests."""
    return [
        _canonical_header(),
        [
            "MILESTONE",
            "M1 Launch",
            None,
            None,
            None,
            None,
            None,
            None,
            WEEK_DATES[0],
            WEEK_DATES[1],
            WEEK_DATES[2],
            WEEK_DATES[3],
        ],
        ["EPIC", "Auth", "Authentication epic", "M1 Launch", None, None, None, None],
        [
            "TASK",
            "Login form",
            "Build login UI",
            "M1 Launch",
            "Auth",
            "High",
            20,
            "Alice",
            10,
            10,
            None,
            None,
        ],
        [
            "TASK",
            "OAuth",
            "Add OAuth provider",
            "M1 Launch",
            "Auth",
            "Medium",
            16,
            "Alice",
            None,
            None,
            8,
            8,
        ],
        ["EPIC", "Payments", "Payments epic", "M1 Launch", None, None, None, None],
        [
            "TASK",
            "Checkout",
            "Checkout flow",
            "M1 Launch",
            "Payments",
            "Critical",
            24,
            "Bob",
            12,
            12,
            None,
            None,
        ],
        [
            "TASK",
            "Refunds",
            "Refund handling",
            "M1 Launch",
            "Payments",
            "Low",
            8,
            "Bob",
            None,
            None,
            4,
            4,
        ],
    ]


@pytest.fixture
def project_with_uploader(db):
    """Project + Developer + User where the user is also the team developer.

    Returns a tuple ``(proj, uploader_user, uploader_dev)``. Most commit
    tests need this base scaffolding. Local variable names avoid the
    module-name shadows that would trip ``from models import user,
    project, ...``.
    """
    now = dt.datetime(2026, 1, 1, 12, 0, 0)
    proj = Project(
        id=1,
        name="P",
        description="",
        status="active",
        key_prefix="PROJ",
        github_repo_urls=[],
        created_at=now,
    )
    db.add(proj)
    uploader_user = User(
        id=10,
        email="uploader@example.com",
        name="Uploader",
        role="developer",
        hashed_password="x",
    )
    db.add(uploader_user)
    uploader_dev = Developer(id=100, name="Uploader", email="uploader@example.com")
    db.add(uploader_dev)
    db.commit()
    return proj, uploader_user, uploader_dev


# ═════════════════════════════════════════════════════════════════════════════
# Section 1 — File format & structure
# ═════════════════════════════════════════════════════════════════════════════


class TestFileFormat:
    def test_happy_path_parses_full_structure(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path, sprint_weeks=2)

        # Output dict has every advertised top-level key
        for key in (
            "meta",
            "tickets",
            "schedule",
            "conflicts",
            "parallel_tasks",
            "availability",
            "warnings",
            "sprints",
            "unscheduled_tasks",
        ):
            assert key in result, f"missing key: {key}"

        # Meta is consistent with the input
        assert result["meta"]["total_tasks"] == 4
        assert result["meta"]["total_assignees"] == 2
        assert result["meta"]["total_weeks"] == 4
        assert result["meta"]["sprint_weeks"] == 2
        assert result["meta"]["week_range"]["start"] == _w(0)
        assert result["meta"]["week_range"]["end"] == _w(3)

    def test_empty_workbook_raises(self, tmp_path):
        wb = openpyxl.Workbook()
        # Strip the default sheet so the workbook genuinely has none. openpyxl
        # forbids removing the last sheet, so we add a placeholder first then
        # remove the default to leave a single (empty) sheet with no rows.
        wb.create_sheet("placeholder")
        del wb["Sheet"]
        # Now remove the placeholder and re-create an empty workbook by
        # writing a workbook with one sheet that has zero rows.
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        # With one empty sheet, no rows means missing required columns → ValueError
        with pytest.raises(ValueError, match="required columns"):
            parse_roadmap(str(path))

    def test_missing_required_columns_raises(self, workbook_factory):
        rows = [
            ["Foo", "Bar", "Baz"],
            ["data", "data", "data"],
        ]
        path = workbook_factory(rows)
        with pytest.raises(ValueError, match="Could not find required columns"):
            parse_roadmap(path)

    def test_no_milestone_row_raises(self, workbook_factory):
        rows = [
            _canonical_header(),
            # No MILESTONE row — only an epic + task, so no week dates anchored
            ["EPIC", "Solo Epic", "", "", None, None, None, None],
            ["TASK", "Orphan task", "", "", "Solo Epic", "Medium", 5, "Alice", 5, None, None, None],
        ]
        path = workbook_factory(rows)
        with pytest.raises(ValueError, match="No MILESTONE row"):
            parse_roadmap(path)

    def test_sheet_picked_by_name_match(self, tmp_path):
        """Workbook with a non-roadmap sheet first, a 'Roadmap' sheet second —
        parser must pick the named one."""
        wb = openpyxl.Workbook()
        first = wb.active
        first.title = "Notes"
        first.append(["random", "garbage"])
        roadmap = wb.create_sheet("Project Roadmap")
        for row in _basic_rows():
            roadmap.append(row)
        path = tmp_path / "named.xlsx"
        wb.save(path)
        result = parse_roadmap(str(path))
        assert result["meta"]["total_tasks"] == 4

    def test_sheet_picked_by_structure_when_no_name_match(self, tmp_path):
        """No sheet contains 'roadmap' — fall back to first sheet with type +
        name columns."""
        wb = openpyxl.Workbook()
        first = wb.active
        first.title = "Garbage"
        first.append(["random", "stuff"])
        data = wb.create_sheet("Plan")
        for row in _basic_rows():
            data.append(row)
        path = tmp_path / "structural.xlsx"
        wb.save(path)
        result = parse_roadmap(str(path))
        assert result["meta"]["total_tasks"] == 4


# ═════════════════════════════════════════════════════════════════════════════
# Section 2 — Column-header detection variations
# ═════════════════════════════════════════════════════════════════════════════


class TestColumnDetection:
    def test_title_alias_for_name(self, workbook_factory):
        rows = _basic_rows()
        rows[0][1] = "Task Title"  # "Name" → "Task Title"
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        assert result["meta"]["total_tasks"] == 4

    def test_owner_alias_for_assignee(self, workbook_factory):
        rows = _basic_rows()
        rows[0][7] = "Owner"  # "Assignee" → "Owner"
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        names = {t["assignee"] for t in result["tickets"]}
        assert names == {"Alice", "Bob"}

    def test_hours_alias_for_effort(self, workbook_factory):
        rows = _basic_rows()
        rows[0][6] = "Hours"  # "Effort (hrs)" → "Hours"
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        # Login form is in row 4 with effort 20
        login = next(t for t in result["tickets"] if t["name"] == "Login form")
        assert login["effort_hrs"] == 20.0

    def test_priority_strict_match(self, workbook_factory):
        """Priority column must literally contain 'priority' or be 'p'/'prio'.
        A column called 'Importance' must NOT be picked up as priority —
        otherwise tasks inherit random text from unrelated columns."""
        rows = _basic_rows()
        rows[0][5] = "Importance"  # ambiguous label
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        # Without a recognised priority column, parser uses the default
        for t in result["tickets"]:
            assert t["priority"] == "Medium"

    def test_priority_short_form(self, workbook_factory):
        rows = _basic_rows()
        rows[0][5] = "P"  # short form for Priority
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        # Now the column IS picked up — Login form keeps its 'High'
        login = next(t for t in result["tickets"] if t["name"] == "Login form")
        assert login["priority"] == "High"


# ═════════════════════════════════════════════════════════════════════════════
# Section 3 — Epic / ticket / milestone extraction
# ═════════════════════════════════════════════════════════════════════════════


class TestExtraction:
    def test_all_tasks_extracted(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path)
        names = {t["name"] for t in result["tickets"]}
        assert names == {"Login form", "OAuth", "Checkout", "Refunds"}

    def test_each_ticket_carries_epic_field(self, workbook_factory):
        """The parser doesn't return a top-level 'epics' list — epic linkage
        lives on each ticket's `epic` key. The commit step uses that to group
        children under their epic parent."""
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path)
        by_name = {t["name"]: t for t in result["tickets"]}
        assert by_name["Login form"]["epic"] == "Auth"
        assert by_name["OAuth"]["epic"] == "Auth"
        assert by_name["Checkout"]["epic"] == "Payments"
        assert by_name["Refunds"]["epic"] == "Payments"

    def test_mixed_case_row_types_accepted(self, workbook_factory):
        rows = _basic_rows()
        rows[1][0] = "milestone"  # lowercase
        rows[2][0] = "Epic"  # title-case
        rows[3][0] = "task"  # lowercase
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        assert result["meta"]["total_tasks"] >= 1

    def test_non_task_rows_skipped(self, workbook_factory):
        """Garbage row types (e.g. 'TOTAL', 'NOTES') must be ignored without
        crashing — only TASK rows become tickets."""
        rows = _basic_rows()
        rows.append(["TOTAL", "All hours", None, None, None, None, 68, None, 22, 22, 12, 12])
        rows.append(["NOTES", "Some note", None, None, None, None, None, None])
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        assert result["meta"]["total_tasks"] == 4  # unchanged

    def test_multi_milestone_uses_correct_week_dates(self, workbook_factory):
        """Two milestones with distinct week sequences — tasks under the
        second milestone must be scheduled against the second milestone's
        weeks, not the first's."""
        m2_weeks = [
            dt.date(2026, 3, 2),
            dt.date(2026, 3, 9),
            dt.date(2026, 3, 16),
            dt.date(2026, 3, 23),
        ]
        rows = _basic_rows()
        rows.append(["MILESTONE", "M2 Polish", None, None, None, None, None, None, *m2_weeks])
        rows.append(["EPIC", "Polish", "Polish epic", "M2 Polish", None, None, None, None])
        rows.append(
            [
                "TASK",
                "UI polish",
                "Tighten visuals",
                "M2 Polish",
                "Polish",
                "Low",
                8,
                "Alice",
                4,
                4,
                None,
                None,
            ]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        polish = next(t for t in result["tickets"] if t["name"] == "UI polish")
        # The two hour-bearing weeks must be M2 weeks 0 and 1
        assert set(polish["active_weeks"]) == {m2_weeks[0].isoformat(), m2_weeks[1].isoformat()}


# ═════════════════════════════════════════════════════════════════════════════
# Section 4 — Sprint computation
# ═════════════════════════════════════════════════════════════════════════════


class TestSprints:
    def test_sprint_count_with_two_week_windows(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path, sprint_weeks=2)
        # 4 weeks / 2-week sprints = 2 sprints
        assert result["meta"]["total_sprints"] == 2
        assert len(result["sprints"]) == 2

    def test_sprint_count_with_three_week_windows(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path, sprint_weeks=3)
        # 4 weeks / 3-week sprints = 2 sprints (3 + 1 trailing)
        assert len(result["sprints"]) == 2
        assert result["sprints"][0]["duration_weeks"] == 3
        assert result["sprints"][1]["duration_weeks"] == 1

    def test_sprint_end_is_friday_of_last_monday(self, workbook_factory):
        """Each week date in the sheet is a Monday; the parser converts the
        last Monday in the sprint window to the Friday of that week as
        end_week. Verify exact dates."""
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path, sprint_weeks=2)
        s1 = result["sprints"][0]
        # Sprint 1: weeks 0–1, end = last Monday + 4 days = 2026-01-12 + 4 = 2026-01-16 (Fri)
        assert s1["start_week"] == _w(0)
        assert s1["end_week"] == "2026-01-16"
        s2 = result["sprints"][1]
        assert s2["start_week"] == _w(2)
        assert s2["end_week"] == "2026-01-30"

    def test_tasks_assigned_to_correct_sprint(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path, sprint_weeks=2)
        s1, s2 = result["sprints"]
        # Login form + Checkout span weeks 0-1 → Sprint 1
        assert "Login form" in s1["tasks"]
        assert "Checkout" in s1["tasks"]
        # OAuth + Refunds span weeks 2-3 → Sprint 2
        assert "OAuth" in s2["tasks"]
        assert "Refunds" in s2["tasks"]

    def test_unscheduled_task_with_no_week_hours(self, workbook_factory):
        rows = _basic_rows()
        # Add an effort-only task with no week hours
        rows.append(
            [
                "TASK",
                "Ghost task",
                "no plan",
                "M1 Launch",
                "Auth",
                "Low",
                4,
                "Alice",
                None,
                None,
                None,
                None,
            ]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path, sprint_weeks=2)
        unscheduled_names = {u["name"] for u in result["unscheduled_tasks"]}
        assert "Ghost task" in unscheduled_names


# ═════════════════════════════════════════════════════════════════════════════
# Section 4b — Multi-milestone date handling (union, gaps, run-aware sprints)
# ═════════════════════════════════════════════════════════════════════════════
#
# Before the union refactor, the parser silently used only the LAST milestone's
# (actually: last task's milestone's) week dates for `meta.week_range` and
# sprint generation. Multi-milestone roadmaps where different milestones had
# different dates dropped most of their tickets into `unscheduled_tasks` with
# no warning. These tests pin the new behaviour:
#
#   - `meta.week_range` covers the UNION of all milestones (min → max).
#   - `meta.missing_weeks` lists calendar gaps that no milestone covered.
#   - Each gap emits a `warnings` entry with issue=`uncovered_week`.
#   - `calculate_sprints` splits the union at every gap into contiguous runs
#     and chunks each run separately — sprints never bridge an uncovered
#     calendar week. Numbering is global across runs (Sprint 1, 2, 3 …).
#   - Tasks under any milestone get scheduled into the sprint(s) whose
#     covered weeks overlap their hours — no silent drops to unscheduled.
#
# Adjacent-date and overlapping-date milestones still produce one contiguous
# run (no gap), so behaviour matches the single-milestone case there.


def _multi_milestone_rows(m1_weeks, m2_weeks, *, m2_task_name="UI polish", m2_hours=(4, 4)):
    """Two-milestone roadmap with one task in each milestone.

    Lets us parametrise the date layout from the test body so we can cover
    adjacent / disjoint / overlapping / gap-in-the-middle scenarios with
    one helper. The two milestones each cover four week-columns from
    `week_col_start` onward; cells beyond `m{n}_weeks`' length are None.
    Each task has hours in its milestone's first two weeks.
    """
    # The header always declares 4 week columns. Each milestone's date row
    # only needs to fill in the cells corresponding to its own weeks; the
    # parser reads `row[week_col_start:]` and stops at the first non-date,
    # so trailing Nones are fine.
    header = _canonical_header()

    def _pad(weeks):
        # Header is 8 left cols + 4 week cols; pad missing weeks with None
        return list(weeks) + [None] * (4 - len(weeks))

    return [
        header,
        ["MILESTONE", "M1", None, None, None, None, None, None, *_pad(m1_weeks)],
        ["EPIC", "Auth", "Authentication", "M1", None, None, None, None],
        # Task under M1: hours in its first two weeks
        [
            "TASK",
            "Login form",
            "",
            "M1",
            "Auth",
            "High",
            20,
            "Alice",
            10,
            10,
            None,
            None,
        ],
        ["MILESTONE", "M2", None, None, None, None, None, None, *_pad(m2_weeks)],
        ["EPIC", "Polish", "UI polish", "M2", None, None, None, None],
        [
            "TASK",
            m2_task_name,
            "",
            "M2",
            "Polish",
            "Low",
            sum(m2_hours),
            "Bob",
            *m2_hours,
            None,
            None,
        ],
    ]


class TestMultiMilestoneUnion:
    def test_same_dates_across_milestones_unchanged_behaviour(self, workbook_factory):
        """Common case: every MILESTONE row shares the same week dates
        (e.g. the standard generated template). Union equals that shared
        sequence; no gaps; sprints identical to the single-milestone era."""
        same = list(WEEK_DATES)  # both milestones use all 4 weeks
        rows = _multi_milestone_rows(same, same)
        path = workbook_factory(rows)
        result = parse_roadmap(path, sprint_weeks=2)
        assert result["meta"]["week_range"]["start"] == _w(0)
        assert result["meta"]["week_range"]["end"] == _w(3)
        assert result["meta"]["missing_weeks"] == []
        # No uncovered_week warnings
        assert not any(w["issue"] == "uncovered_week" for w in result["warnings"])
        # 2 sprints, one for each pair of weeks
        assert result["meta"]["total_sprints"] == 2

    def test_adjacent_disjoint_milestones_form_single_run(self, workbook_factory):
        """M1 covers W1-W2, M2 covers W3-W4. Union is contiguous (no gap)
        so it's one run, chunked into 2 sprints. Each task lands in the
        sprint whose calendar weeks match its milestone."""
        m1 = [WEEK_DATES[0], WEEK_DATES[1]]
        m2 = [WEEK_DATES[2], WEEK_DATES[3]]
        rows = _multi_milestone_rows(m1, m2, m2_hours=(5, 5))
        path = workbook_factory(rows)
        result = parse_roadmap(path, sprint_weeks=2)
        # No gap → no warning
        assert result["meta"]["missing_weeks"] == []
        # Two sprints: one for M1's weeks, one for M2's
        assert len(result["sprints"]) == 2
        s1, s2 = result["sprints"]
        assert s1["start_week"] == _w(0)
        assert s1["end_week"] == "2026-01-16"
        assert s2["start_week"] == _w(2)
        assert s2["end_week"] == "2026-01-30"
        # Each milestone's task in its calendar-correct sprint
        assert "Login form" in s1["tasks"]
        assert "UI polish" in s2["tasks"]
        assert result["unscheduled_tasks"] == []

    def test_milestones_with_gap_emit_warnings_and_split_runs(self, workbook_factory):
        """The headline scenario: M1 covers W1-W2 (Jan 5-12), M2 jumps to
        2026-02-09 (skipping Jan 19, Jan 26, Feb 2). Union = [Jan 5, Jan 12,
        Feb 9, Feb 16]. Missing = [Jan 19, Jan 26, Feb 2]. Three warnings,
        two sprints (one per run), no sprint bridging the gap."""
        m1 = [WEEK_DATES[0], WEEK_DATES[1]]  # Jan 5, Jan 12
        m2 = [dt.date(2026, 2, 9), dt.date(2026, 2, 16)]  # Feb 9, Feb 16
        rows = _multi_milestone_rows(m1, m2, m2_hours=(6, 6))
        path = workbook_factory(rows)
        result = parse_roadmap(path, sprint_weeks=2)

        # meta exposes the gaps
        assert result["meta"]["missing_weeks"] == ["2026-01-19", "2026-01-26", "2026-02-02"]
        # week_range spans full union
        assert result["meta"]["week_range"]["start"] == _w(0)
        assert result["meta"]["week_range"]["end"] == "2026-02-16"

        # Each gap surfaces a warning
        uncovered = [w for w in result["warnings"] if w["issue"] == "uncovered_week"]
        assert {w["detail"].split()[2] for w in uncovered} == {
            "2026-01-19",
            "2026-01-26",
            "2026-02-02",
        }

        # Sprints: one per run, calendar-continuous within each
        assert len(result["sprints"]) == 2
        s1, s2 = result["sprints"]
        # Sprint 1: M1's run
        assert s1["start_week"] == "2026-01-05"
        assert s1["end_week"] == "2026-01-16"
        assert s1["week_dates"] == ["2026-01-05", "2026-01-12"]
        # Sprint 2: M2's run — starts on Feb 9, NOT bridging from Jan
        assert s2["start_week"] == "2026-02-09"
        assert s2["end_week"] == "2026-02-20"
        assert s2["week_dates"] == ["2026-02-09", "2026-02-16"]

        # Both tasks make it into a sprint — nothing silently dropped
        assert "Login form" in s1["tasks"]
        assert "UI polish" in s2["tasks"]
        assert result["unscheduled_tasks"] == []

    def test_overlapping_milestone_dates_collapse_to_single_run(self, workbook_factory):
        """M1 covers W1-W2, M2 covers W2-W3. Their overlap on W2 collapses
        in the union (sets dedupe). One contiguous run [W1, W2, W3]."""
        m1 = [WEEK_DATES[0], WEEK_DATES[1]]
        m2 = [WEEK_DATES[1], WEEK_DATES[2]]
        rows = _multi_milestone_rows(m1, m2, m2_hours=(4, 4))
        path = workbook_factory(rows)
        result = parse_roadmap(path, sprint_weeks=2)
        assert result["meta"]["missing_weeks"] == []
        assert result["meta"]["total_weeks"] == 3  # union dedupes the shared W2
        # sprint_weeks=2 over a 3-week run = Sprint 1 [W1, W2], Sprint 2 [W3]
        assert len(result["sprints"]) == 2
        assert result["sprints"][0]["duration_weeks"] == 2
        assert result["sprints"][1]["duration_weeks"] == 1

    def test_gap_falling_mid_chunk_breaks_sprint_at_run_boundary(self, workbook_factory):
        """Critical correctness case: a gap that lands inside what would
        otherwise be a sprint chunk must split the sprint, not bridge across.

        Union [W1, W2, W3, W5, W6] with sprint_weeks=2 must produce:
            Sprint 1: [W1, W2]
            Sprint 2: [W3]          ← runt; closes M1's run at W4 gap
            Sprint 3: [W5, W6]
        and NOT:
            Sprint 2: [W3, W5]      ← bridges Jan 19 → Feb 2 (BAD)
        """
        # W1=Jan 5, W2=Jan 12, W3=Jan 19, (W4=Jan 26 missing), W5=Feb 2, W6=Feb 9
        m1 = [
            dt.date(2026, 1, 5),
            dt.date(2026, 1, 12),
            dt.date(2026, 1, 19),
        ]
        m2 = [dt.date(2026, 2, 2), dt.date(2026, 2, 9)]
        rows = _multi_milestone_rows(m1, m2, m2_hours=(8, 8))
        path = workbook_factory(rows)
        result = parse_roadmap(path, sprint_weeks=2)

        # Gap detection: only Jan 26 is missing
        assert result["meta"]["missing_weeks"] == ["2026-01-26"]

        # Three sprints — the runt Sprint 2 prevents the bridge
        assert len(result["sprints"]) == 3
        s1, s2, s3 = result["sprints"]
        assert s1["week_dates"] == ["2026-01-05", "2026-01-12"]
        assert s2["week_dates"] == ["2026-01-19"]
        assert s2["duration_weeks"] == 1
        # Sprint 2 must NOT include Feb 2 — the calendar-bridging bug
        assert "2026-02-02" not in s2["week_dates"]
        assert s3["week_dates"] == ["2026-02-02", "2026-02-09"]
        # Sprints are numbered globally across runs (1, 2, 3)
        assert [s["number"] for s in result["sprints"]] == [1, 2, 3]

    def test_pre_union_bug_regression_m2_task_not_silently_unscheduled(self, workbook_factory):
        """Regression guard for the multi-milestone bug. Before the union
        refactor, sprints were built from only one milestone's weeks, so a
        task under the OTHER milestone landed in `unscheduled_tasks` with
        reason 'hours outside sprint boundaries'. The fix is verified by
        checking that no task from either milestone ends up unscheduled
        purely because of a milestone mismatch."""
        m1 = [WEEK_DATES[0], WEEK_DATES[1]]
        m2 = [dt.date(2026, 3, 2), dt.date(2026, 3, 9)]
        rows = _multi_milestone_rows(m1, m2)
        path = workbook_factory(rows)
        result = parse_roadmap(path, sprint_weeks=2)
        unscheduled_names = {u["name"] for u in result["unscheduled_tasks"]}
        assert "Login form" not in unscheduled_names
        assert "UI polish" not in unscheduled_names
        # And UI polish actually lands in the M2-run sprint
        m2_sprint = next(s for s in result["sprints"] if s["start_week"] == "2026-03-02")
        assert "UI polish" in m2_sprint["tasks"]


# ═════════════════════════════════════════════════════════════════════════════
# Section 5 — Hours: per-week schedule, conflicts, parallel tasks
# ═════════════════════════════════════════════════════════════════════════════


class TestHoursAndConflicts:
    def test_schedule_per_assignee_per_week(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path)
        # Alice has 10h in W0 (Login form)
        assert result["schedule"]["Alice"][_w(0)]["total_hrs"] == 10
        assert "Login form" in result["schedule"]["Alice"][_w(0)]["tasks"]
        # Bob has 12h in W0 (Checkout)
        assert result["schedule"]["Bob"][_w(0)]["total_hrs"] == 12

    def test_conflict_when_same_assignee_two_tasks_one_week(self, workbook_factory):
        rows = _basic_rows()
        # Stack a second task on Alice in W0 to manufacture a conflict
        rows.append(
            [
                "TASK",
                "Side quest",
                "extra work",
                "M1 Launch",
                "Auth",
                "Medium",
                5,
                "Alice",
                5,
                None,
                None,
                None,
            ]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        alice_w0 = [
            c for c in result["conflicts"] if c["assignee"] == "Alice" and c["week"] == _w(0)
        ]
        assert len(alice_w0) == 1
        # 10 + 5 = 15h, under the 40h threshold → not overbooked
        assert alice_w0[0]["total_hrs"] == 15
        assert alice_w0[0]["overbooked"] is False
        assert set(alice_w0[0]["tasks"]) == {"Login form", "Side quest"}

    def test_conflict_overbooked_above_threshold(self, workbook_factory):
        rows = [
            _canonical_header(),
            [
                "MILESTONE",
                "M1",
                None,
                None,
                None,
                None,
                None,
                None,
                WEEK_DATES[0],
                WEEK_DATES[1],
                WEEK_DATES[2],
                WEEK_DATES[3],
            ],
            ["EPIC", "E1", "", "M1", None, None, None, None],
            ["TASK", "Big A", "", "M1", "E1", "High", 30, "Alice", 30, None, None, None],
            ["TASK", "Big B", "", "M1", "E1", "High", 20, "Alice", 20, None, None, None],
        ]
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        conflicts = [c for c in result["conflicts"] if c["assignee"] == "Alice"]
        assert len(conflicts) == 1
        assert conflicts[0]["total_hrs"] == 50
        assert conflicts[0]["overbooked"] is True  # 50 > 40

    def test_parallel_tasks_different_assignees_same_week(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path)
        # Login form (Alice) and Checkout (Bob) both have hours in W0
        pairs = result["parallel_tasks"]
        w0_pair = next(
            (
                p
                for p in pairs
                if p["week"] == _w(0) and {p["task_a"], p["task_b"]} == {"Login form", "Checkout"}
            ),
            None,
        )
        assert w0_pair is not None
        assert {w0_pair["assignee_a"], w0_pair["assignee_b"]} == {"Alice", "Bob"}

    def test_availability_first_free_week(self, workbook_factory):
        path = workbook_factory(_basic_rows())
        result = parse_roadmap(path)
        # Alice's last busy week is W3 (OAuth has 8h in W3); first_free = W3 + 7d
        alice = result["availability"]["Alice"]
        assert alice["last_busy_week"] == _w(3)
        assert alice["first_free_week"] == "2026-02-02"  # 2026-01-26 + 7d


# ═════════════════════════════════════════════════════════════════════════════
# Section 6 — Warnings
# ═════════════════════════════════════════════════════════════════════════════


class TestWarnings:
    def test_unassigned_task_emits_warning(self, workbook_factory):
        rows = _basic_rows()
        rows.append(
            ["TASK", "Orphan", "", "M1 Launch", "Auth", "Medium", 5, None, 5, None, None, None]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        unassigned = [
            w for w in result["warnings"] if w["issue"] == "unassigned" and w["task"] == "Orphan"
        ]
        assert len(unassigned) == 1

    def test_effort_mismatch_warning(self, workbook_factory):
        rows = _basic_rows()
        # Effort col says 20 but weeks sum to 30 (diff 10 > tolerance)
        rows.append(
            ["TASK", "Mismatched", "", "M1 Launch", "Auth", "High", 20, "Alice", 15, 15, None, None]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        mismatches = [
            w
            for w in result["warnings"]
            if w["issue"] == "effort_mismatch" and w["task"] == "Mismatched"
        ]
        assert len(mismatches) == 1

    def test_effort_within_tolerance_no_warning(self, workbook_factory):
        rows = _basic_rows()
        # 10 vs 10.3 — diff 0.3 ≤ 0.5 tolerance → no warning
        rows.append(
            [
                "TASK",
                "Close enough",
                "",
                "M1 Launch",
                "Auth",
                "Low",
                10,
                "Alice",
                10.3,
                None,
                None,
                None,
            ]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        assert not any(
            w["task"] == "Close enough" and w["issue"] == "effort_mismatch"
            for w in result["warnings"]
        )

    def test_no_weeks_planned_warning(self, workbook_factory):
        rows = _basic_rows()
        rows.append(
            [
                "TASK",
                "Effort-only",
                "",
                "M1 Launch",
                "Auth",
                "Medium",
                12,
                "Alice",
                None,
                None,
                None,
                None,
            ]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        ghost = [
            w
            for w in result["warnings"]
            if w["issue"] == "no_weeks_planned" and w["task"] == "Effort-only"
        ]
        assert len(ghost) == 1

    def test_invalid_effort_cell_parses_to_none(self, workbook_factory):
        """A non-numeric effort value (e.g. 'TBD') must not crash; the field
        becomes None and the task is otherwise processed normally."""
        rows = _basic_rows()
        rows.append(
            [
                "TASK",
                "Fuzzy effort",
                "",
                "M1 Launch",
                "Auth",
                "Medium",
                "TBD",
                "Alice",
                4,
                None,
                None,
                None,
            ]
        )
        path = workbook_factory(rows)
        result = parse_roadmap(path)
        fuzzy = next(t for t in result["tickets"] if t["name"] == "Fuzzy effort")
        assert fuzzy["effort_hrs"] is None
        # And no effort_mismatch warning (effort is None)
        assert not any(
            w["task"] == "Fuzzy effort" and w["issue"] == "effort_mismatch"
            for w in result["warnings"]
        )


# ═════════════════════════════════════════════════════════════════════════════
# Section 7 — Commit: end-to-end DB materialisation
# ═════════════════════════════════════════════════════════════════════════════


def _commit_from_workbook(workbook_factory, db, proj, uploader, *, sprint_weeks=2, rows=None):
    """Parse a workbook and run the commit endpoint inline. Returns the
    response dict so individual tests can assert on it."""
    rows = rows or _basic_rows()
    path = workbook_factory(rows)
    parsed = parse_roadmap(path, sprint_weeks=sprint_weeks)
    request = RoadmapCommitRequest(project_id=proj.id, parsed_data=parsed)
    return commit_roadmap_tickets(request=request, db=db, current_user=uploader)


class TestCommitHappyPath:
    def test_creates_expected_count_of_epics_tasks_sprints(
        self, workbook_factory, db, project_with_uploader
    ):
        proj, uploader, _ = project_with_uploader
        response = _commit_from_workbook(workbook_factory, db, proj, uploader)
        assert response["epics_created"] == 2
        assert response["tickets_created"] == 4
        assert response["sprints_created"] == 2

    def test_response_message_pluralisation(self, workbook_factory, db, project_with_uploader):
        proj, uploader, _ = project_with_uploader
        response = _commit_from_workbook(workbook_factory, db, proj, uploader)
        assert "2 epics" in response["message"]
        assert "4 tasks" in response["message"]
        assert "2 sprints" in response["message"]

    def test_epics_persisted_with_correct_type_and_title(
        self, workbook_factory, db, project_with_uploader
    ):
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        epics = (
            db.query(WorkItem).filter(WorkItem.project_id == proj.id, WorkItem.type == "epic").all()
        )
        titles = {e.title for e in epics}
        assert titles == {"Auth", "Payments"}
        for e in epics:
            assert e.status == "todo"
            # Epics are created with no assignee — they're aggregation nodes
            assert e.assignee_id is None

    def test_tasks_persisted_as_user_stories(self, workbook_factory, db, project_with_uploader):
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        stories = (
            db.query(WorkItem)
            .filter(WorkItem.project_id == proj.id, WorkItem.type == "user_story")
            .all()
        )
        titles = {s.title for s in stories}
        assert titles == {"Login form", "OAuth", "Checkout", "Refunds"}

    def test_reporter_is_the_uploader(self, workbook_factory, db, project_with_uploader):
        proj, uploader, dev = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        items = db.query(WorkItem).filter(WorkItem.project_id == proj.id).all()
        # Every item — epic OR task — is attributed to the uploader
        for item in items:
            assert item.reporter_id == dev.id, (
                f"{item.title} reporter_id should be uploader's dev id"
            )


class TestCommitHierarchy:
    """The user's explicit ask: 'do child items belong to parent as shown in
    file'. Each story's `epic_id` must point to the epic with the matching
    name from the xlsx."""

    def test_each_task_links_to_correct_parent_epic(
        self, workbook_factory, db, project_with_uploader
    ):
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        epic_by_name = {
            e.title: e
            for e in db.query(WorkItem)
            .filter(WorkItem.project_id == proj.id, WorkItem.type == "epic")
            .all()
        }
        story_by_name = {
            s.title: s
            for s in db.query(WorkItem)
            .filter(WorkItem.project_id == proj.id, WorkItem.type == "user_story")
            .all()
        }
        # Mirror the rows in _basic_rows():
        assert story_by_name["Login form"].epic_id == epic_by_name["Auth"].id
        assert story_by_name["OAuth"].epic_id == epic_by_name["Auth"].id
        assert story_by_name["Checkout"].epic_id == epic_by_name["Payments"].id
        assert story_by_name["Refunds"].epic_id == epic_by_name["Payments"].id

    def test_epics_deduped_across_tickets(self, workbook_factory, db, project_with_uploader):
        """Three tickets all under 'Auth' must produce ONE epic row, not three."""
        proj, uploader, _ = project_with_uploader
        rows = _basic_rows()
        # Add a third task under Auth
        rows.append(
            ["TASK", "Logout", "", "M1 Launch", "Auth", "Low", 4, "Alice", 4, None, None, None]
        )
        _commit_from_workbook(workbook_factory, db, proj, uploader, rows=rows)
        auth_epics = (
            db.query(WorkItem)
            .filter(
                WorkItem.project_id == proj.id,
                WorkItem.type == "epic",
                WorkItem.title == "Auth",
            )
            .all()
        )
        assert len(auth_epics) == 1

    def test_no_orphan_tasks(self, workbook_factory, db, project_with_uploader):
        """Every imported story has a non-null epic_id — nothing is orphaned."""
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        orphans = (
            db.query(WorkItem)
            .filter(
                WorkItem.project_id == proj.id,
                WorkItem.type == "user_story",
                WorkItem.epic_id.is_(None),
            )
            .all()
        )
        assert orphans == []


class TestCommitHoursRollup:
    """The user's other explicit ask: 'do child items hours add up to parent
    hours'. After commit, epic.estimated_hours must equal the sum of its
    children's estimated_hours."""

    def test_epic_hours_sum_to_children(self, workbook_factory, db, project_with_uploader):
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        auth = (
            db.query(WorkItem)
            .filter(
                WorkItem.project_id == proj.id,
                WorkItem.type == "epic",
                WorkItem.title == "Auth",
            )
            .one()
        )
        payments = (
            db.query(WorkItem)
            .filter(
                WorkItem.project_id == proj.id,
                WorkItem.type == "epic",
                WorkItem.title == "Payments",
            )
            .one()
        )
        # Auth = Login form (20) + OAuth (16) = 36
        assert auth.estimated_hours == 36
        # Payments = Checkout (24) + Refunds (8) = 32
        assert payments.estimated_hours == 32

    def test_each_task_estimated_hours_matches_file(
        self, workbook_factory, db, project_with_uploader
    ):
        """Per-task fidelity. Effort 8.5 should land as 8 (int truncation in
        create_work_item)."""
        proj, uploader, _ = project_with_uploader
        rows = _basic_rows()
        # Add a decimal-effort task
        rows.append(
            [
                "TASK",
                "Tweak copy",
                "",
                "M1 Launch",
                "Auth",
                "Low",
                8.5,
                "Alice",
                8.5,
                None,
                None,
                None,
            ]
        )
        _commit_from_workbook(workbook_factory, db, proj, uploader, rows=rows)
        tweak = db.query(WorkItem).filter(WorkItem.title == "Tweak copy").one()
        assert tweak.estimated_hours == 8  # int truncation
        assert tweak.remaining_hours == 8

    def test_remaining_hours_initialised_to_estimate(
        self, workbook_factory, db, project_with_uploader
    ):
        """A newly imported task hasn't logged any work yet, so
        remaining_hours starts equal to estimated_hours."""
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        stories = (
            db.query(WorkItem)
            .filter(
                WorkItem.project_id == proj.id,
                WorkItem.type == "user_story",
            )
            .all()
        )
        for s in stories:
            assert s.remaining_hours == s.estimated_hours


class TestCommitSprints:
    def test_sprints_persisted_with_correct_dates(
        self, workbook_factory, db, project_with_uploader
    ):
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        sprints = (
            db.query(Sprint).filter(Sprint.project_id == proj.id).order_by(Sprint.start_date).all()
        )
        assert [s.name for s in sprints] == ["Sprint 1", "Sprint 2"]
        # Sprint 1: 2026-01-05 → 2026-01-16 (Fri of last Mon in window)
        assert sprints[0].start_date.date() == dt.date(2026, 1, 5)
        assert sprints[0].end_date.date() == dt.date(2026, 1, 16)
        assert sprints[0].status == "planned"

    def test_tasks_get_sprint_id_and_due_date(self, workbook_factory, db, project_with_uploader):
        proj, uploader, _ = project_with_uploader
        response = _commit_from_workbook(workbook_factory, db, proj, uploader)
        assert response["tasks_assigned_to_sprints"] == 4

        sprints = {s.name: s for s in db.query(Sprint).all()}
        login = db.query(WorkItem).filter(WorkItem.title == "Login form").one()
        # Login form spans W0-W1 → Sprint 1
        assert login.sprint_id == sprints["Sprint 1"].id
        # Due date = sprint end (2026-01-16)
        assert login.due_date.date() == sprints["Sprint 1"].end_date.date()


class TestCommitAssigneeResolution:
    def test_known_assignee_resolved_by_name(self, workbook_factory, db, project_with_uploader):
        proj, uploader, _ = project_with_uploader
        alice = Developer(id=200, name="Alice", email="alice@example.com")
        bob = Developer(id=201, name="Bob", email="bob@example.com")
        db.add_all([alice, bob])
        db.commit()
        response = _commit_from_workbook(workbook_factory, db, proj, uploader)
        assert response["assignees_not_found"] == 0
        login = db.query(WorkItem).filter(WorkItem.title == "Login form").one()
        assert login.assignee_id == alice.id
        checkout = db.query(WorkItem).filter(WorkItem.title == "Checkout").one()
        assert checkout.assignee_id == bob.id

    def test_unknown_assignee_falls_back_to_uploader(
        self, workbook_factory, db, project_with_uploader
    ):
        """No Alice/Bob in the developer table — every task must fall back to
        the uploader and increment `assignees_not_found`."""
        proj, uploader, dev = project_with_uploader
        response = _commit_from_workbook(workbook_factory, db, proj, uploader)
        # 4 tasks, none of their assignees exist as developers
        assert response["assignees_not_found"] == 4
        stories = (
            db.query(WorkItem)
            .filter(WorkItem.project_id == proj.id, WorkItem.type == "user_story")
            .all()
        )
        for s in stories:
            assert s.assignee_id == dev.id  # uploader fallback

    def test_blank_assignee_defaults_to_uploader_without_counting(
        self, workbook_factory, db, project_with_uploader
    ):
        """A blank assignee field is data quality, not 'name not found' —
        it does NOT increment the not-found counter (the counter only fires
        when a name was given but didn't match)."""
        proj, uploader, dev = project_with_uploader
        rows = [
            _canonical_header(),
            [
                "MILESTONE",
                "M1",
                None,
                None,
                None,
                None,
                None,
                None,
                WEEK_DATES[0],
                WEEK_DATES[1],
                WEEK_DATES[2],
                WEEK_DATES[3],
            ],
            ["EPIC", "E1", "", "M1", None, None, None, None],
            ["TASK", "Solo", "", "M1", "E1", "Medium", 5, None, 5, None, None, None],
        ]
        response = _commit_from_workbook(workbook_factory, db, proj, uploader, rows=rows)
        assert response["assignees_not_found"] == 0
        solo = db.query(WorkItem).filter(WorkItem.title == "Solo").one()
        assert solo.assignee_id == dev.id


class TestCommitActivityLog:
    def test_single_activity_log_entry_created(self, workbook_factory, db, project_with_uploader):
        proj, uploader, _ = project_with_uploader
        _commit_from_workbook(workbook_factory, db, proj, uploader)
        logs = (
            db.query(ActivityLog)
            .filter(
                ActivityLog.project_id == proj.id,
                ActivityLog.entity_type == "roadmap",
            )
            .all()
        )
        assert len(logs) == 1
        log = logs[0]
        assert log.action == "created"
        assert log.user_id == uploader.id
        assert log.details["epics_created"] == 2
        assert log.details["tasks_created"] == 4
        assert log.details["sprints_created"] == 2
        # The headline title encodes the counts for the activity feed
        assert "2 epics" in log.title
        assert "4 tasks" in log.title


class TestCommitErrors:
    def test_unknown_project_id_raises_404(self, db, project_with_uploader):
        _, uploader, _ = project_with_uploader
        request = RoadmapCommitRequest(project_id=9999, parsed_data={"tickets": []})
        from fastapi import HTTPException

        with pytest.raises(HTTPException) as excinfo:
            commit_roadmap_tickets(request=request, db=db, current_user=uploader)
        assert excinfo.value.status_code == 404

    def test_user_without_developer_row_returns_500_via_wrap(self, db):
        """When the calling uploader has no Developer row we treat them as 'not a
        team member' (400). The endpoint's outer try/except catches the
        HTTPException and re-raises it as a 500 — that's the current,
        intentional behaviour and these tests pin it so future refactors are
        explicit."""
        proj = Project(
            id=2,
            name="Q",
            description="",
            status="active",
            key_prefix="QQQ",
            github_repo_urls=[],
            created_at=dt.datetime(2026, 1, 1),
        )
        uploader = User(
            id=20, email="ghost@example.com", name="Ghost", role="developer", hashed_password="x"
        )
        db.add_all([proj, uploader])
        db.commit()
        request = RoadmapCommitRequest(project_id=proj.id, parsed_data={"tickets": []})
        from fastapi import HTTPException

        with pytest.raises(HTTPException) as excinfo:
            commit_roadmap_tickets(request=request, db=db, current_user=uploader)
        # Outer wrap rebrands the inner 400 as 500, with the original detail
        # surfaced in the message.
        assert excinfo.value.status_code == 500
        assert "team member" in str(excinfo.value.detail)


# ═════════════════════════════════════════════════════════════════════════════
# Section 8 — Full end-to-end: file → parse → commit → DB
# ═════════════════════════════════════════════════════════════════════════════


class TestEndToEnd:
    """One realistic xlsx exercising every behaviour: multi-milestone,
    decimal effort, unknown assignee, overbook conflict, unscheduled task."""

    def _build_realistic_workbook(self, workbook_factory):
        return workbook_factory(
            [
                _canonical_header(),
                [
                    "MILESTONE",
                    "M1 Launch",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    WEEK_DATES[0],
                    WEEK_DATES[1],
                    WEEK_DATES[2],
                    WEEK_DATES[3],
                ],
                ["EPIC", "Auth", "Authentication", "M1 Launch", None, None, None, None],
                [
                    "TASK",
                    "Login form",
                    "",
                    "M1 Launch",
                    "Auth",
                    "High",
                    20,
                    "Alice",
                    10,
                    10,
                    None,
                    None,
                ],
                [
                    "TASK",
                    "Password reset",
                    "",
                    "M1 Launch",
                    "Auth",
                    "Medium",
                    35,
                    "Alice",
                    35,
                    None,
                    None,
                    None,
                ],  # 10 + 35 = 45h W0 → overbooked
                ["EPIC", "Billing", "", "M1 Launch", None, None, None, None],
                [
                    "TASK",
                    "Invoice gen",
                    "",
                    "M1 Launch",
                    "Billing",
                    "Critical",
                    8,
                    "NobodyMatches",
                    4,
                    4,
                    None,
                    None,
                ],  # unknown assignee
                [
                    "TASK",
                    "Ghost",
                    "",
                    "M1 Launch",
                    "Billing",
                    "Low",
                    4,
                    "Alice",
                    None,
                    None,
                    None,
                    None,
                ],  # unscheduled
            ]
        )

    def test_full_pipeline_state(self, workbook_factory, db, project_with_uploader):
        proj, uploader, dev = project_with_uploader
        alice = Developer(id=200, name="Alice", email="alice@example.com")
        db.add(alice)
        db.commit()

        path = self._build_realistic_workbook(workbook_factory)
        parsed = parse_roadmap(path, sprint_weeks=2)

        # ── Parser side ────────────────────────────────────────────────
        # Overbook conflict on Alice W0 (10 + 35 = 45 > 40)
        alice_overbook = [
            c
            for c in parsed["conflicts"]
            if c["assignee"] == "Alice" and c["week"] == _w(0) and c["overbooked"]
        ]
        assert len(alice_overbook) == 1
        # Ghost is unscheduled
        ghost = [u for u in parsed["unscheduled_tasks"] if u["name"] == "Ghost"]
        assert len(ghost) == 1

        # ── Commit side ────────────────────────────────────────────────
        request = RoadmapCommitRequest(project_id=proj.id, parsed_data=parsed)
        response = commit_roadmap_tickets(request=request, db=db, current_user=uploader)

        assert response["epics_created"] == 2
        assert response["tickets_created"] == 4
        assert response["assignees_not_found"] == 1  # NobodyMatches only

        # Auth epic hours = Login (20) + PW reset (35) = 55
        auth = db.query(WorkItem).filter(WorkItem.title == "Auth").one()
        assert auth.estimated_hours == 55
        # Billing epic hours = Invoice (8) + Ghost (4) = 12
        billing = db.query(WorkItem).filter(WorkItem.title == "Billing").one()
        assert billing.estimated_hours == 12

        # Invoice gen fell back to uploader
        invoice = db.query(WorkItem).filter(WorkItem.title == "Invoice gen").one()
        assert invoice.assignee_id == dev.id

        # Alice tasks point at Alice's dev id
        login = db.query(WorkItem).filter(WorkItem.title == "Login form").one()
        assert login.assignee_id == alice.id

        # Activity log captures the import
        log = (
            db.query(ActivityLog)
            .filter(
                ActivityLog.project_id == proj.id,
                ActivityLog.entity_type == "roadmap",
            )
            .one()
        )
        assert log.details["assignees_not_found"] == 1
