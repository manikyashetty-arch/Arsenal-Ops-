"""
parse_roadmap.py
----------------
Parses the AAI Ops roadmap Excel file and outputs:
  - tickets        : list of all TASK rows as structured dicts
  - schedule       : per-assignee, per-week hours breakdown
  - conflicts      : same assignee working >40 hrs in a week
  - parallel_tasks : tasks running in same week (different assignees) — fine
  - availability   : first free week per assignee after their last task
  - warnings       : unassigned tasks, effort mismatches

Usage:
    python parse_roadmap.py                          # prints JSON to stdout
    python parse_roadmap.py --file my_roadmap.xlsx   # custom file path
"""

import sys
import json
import datetime
import argparse
import openpyxl
from logging_config import setup_logger

logger = setup_logger("parser")

# ── Constants ─────────────────────────────────────────────────────────────────
HOURS_PER_WEEK = 40         # threshold for "fully booked"


# ── Helpers ───────────────────────────────────────────────────────────────────

def col_index_to_letter(idx):
    """1-based column index → Excel letter (e.g. 9 → 'I')."""
    return openpyxl.utils.get_column_letter(idx)


def parse_effort(raw):
    """Convert effort cell value to float, or None if blank/invalid."""
    if raw is None:
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def date_key(dt):
    """datetime → 'YYYY-MM-DD' string used as dict key."""
    if isinstance(dt, datetime.datetime):
        return dt.date().isoformat()
    if isinstance(dt, datetime.date):
        return dt.isoformat()
    return str(dt)


# ── Core parser ───────────────────────────────────────────────────────────────

def parse(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # ── Find the sheet with roadmap data (support multiple sheets) ─────────────
    # Try all sheets to find one with the required structure
    ws = None
    sheet_names = wb.sheetnames
    
    if not sheet_names:
        raise ValueError("Excel file has no sheets")
    
    # First, try to find a sheet with "roadmap" in the name
    for sheet_name in sheet_names:
        if "roadmap" in sheet_name.lower():
            ws = wb[sheet_name]
            break
    
    # If not found, try all sheets and use the first one with valid structure
    if ws is None:
        for sheet_name in sheet_names:
            candidate_ws = wb[sheet_name]
            candidate_rows = list(candidate_ws.iter_rows(min_row=1, values_only=True))
            
            if candidate_rows:
                header_row = candidate_rows[0]
                # Check if this sheet has the required columns (type and name)
                header_lower = [str(h).strip().lower() if h else "" for h in header_row]
                has_type = any("type" in h for h in header_lower)
                has_name = any("name" in h or "title" in h for h in header_lower)
                
                if has_type and has_name:
                    ws = candidate_ws
                    break
    
    # Fallback: use first sheet if no valid structure found
    if ws is None:
        ws = wb.active
        logger.warning(f"Could not identify roadmap sheet, using default: {ws.title}")
    
    logger.info(f"Using sheet: {ws.title}")

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    
    # ── Step 0: Build column mapping from header row ────────────────────────────
    # Find columns by name instead of hard-coded positions for flexibility
    header_row = all_rows[0] if all_rows else []
    
    # Define what column names we're looking for (case-insensitive)
    column_names_to_find = {
        "type": ["type", "row type", "item type", "task type"],
        "name": ["name", "title", "task name", "task title"],
        "description": ["description", "details", "desc"],
        "milestone": ["milestone", "phase", "release"],
        "epic": ["epic", "feature", "epic name"],
        "priority": ["priority", "p", "prio"],
        "effort_hrs": ["effort", "effort_hrs", "effort hours", "hours", "hrs", "estimate"],
        "assignee": ["assignee", "assigned to", "owner", "developer", "dev"],
    }
    
    # Build column index mapping
    col_mapping = {}
    for field_name, possible_names in column_names_to_find.items():
        for col_idx, header_cell in enumerate(header_row):
            if header_cell is None:
                continue
            header_lower = str(header_cell).strip().lower()
            # For priority, be VERY specific - must contain "priority" or be exactly "p" or "prio"
            if field_name == "priority":
                # Priority: be more strict, only match if "priority" is a standalone word
                if "priority" in header_lower or header_lower in ["p", "prio"]:
                    col_mapping[field_name] = col_idx
                    break
            else:
                if any(name.lower() in header_lower for name in possible_names):
                    col_mapping[field_name] = col_idx
                    break
    
    # Debug: Show header row
    logger.debug(f"═══ HEADER ROW ANALYSIS ═══")
    logger.debug(f"Total columns: {len(header_row)}")
    logger.debug(f"All Column Headers:")
    for idx, h in enumerate(header_row):
        col_letter = col_index_to_letter(idx + 1)
        logger.debug(f"  [{col_letter}] (idx {idx}): '{h}'")
    
    # Verify we found all critical columns
    required_cols = ["type", "name"]
    missing = [c for c in required_cols if c not in col_mapping]
    if missing:
        raise ValueError(
            f"Could not find required columns: {missing}. "
            f"Available columns: {[str(h) for h in header_row if h]}"
        )
    
    # Use default indices if columns not found (fallback to old behavior)
    # BUT: Only for columns that make sense to have defaults
    # For optional columns like priority, don't use fallback indices - use None instead
    if "description" not in col_mapping:
        col_mapping["description"] = 2
    if "milestone" not in col_mapping:
        col_mapping["milestone"] = 3
    if "epic" not in col_mapping:
        col_mapping["epic"] = 4
    # NOTE: priority, effort_hrs, assignee - if not found, leave out of col_mapping
    # Don't use fallback indices for these as they might point to wrong columns
    if "effort_hrs" not in col_mapping:
        col_mapping["effort_hrs"] = 6
    if "assignee" not in col_mapping:
        col_mapping["assignee"] = 7
    
    # Debug: log the column mapping
    logger.debug(f"═══ IDENTIFIED COLUMNS ═══")
    for field, col_idx in sorted(col_mapping.items(), key=lambda x: x[1]):
        col_letter = col_index_to_letter(col_idx + 1)  # Convert to 1-based for display
        header_val = header_row[col_idx] if col_idx < len(header_row) else "?"
        status = "✓ Found" if header_row[col_idx] else "⚠ Default"
        logger.debug(f"  {field:15} → [{col_letter}] idx {col_idx:2d} = '{header_val}' {status}")
    
    # Find the first week column by looking at the MILESTONE row (which has actual dates)
    # This is more reliable than looking at headers since week columns might not have headers
    week_col_start = max(col_mapping.values()) + 1
    
    # Scan MILESTONE rows to find where the week dates start
    for row in all_rows[1:]:  # Skip header row
        row_type = row[col_mapping["type"]]
        row_type = str(row_type).strip().upper() if row_type else ""
        
        if row_type == "MILESTONE":
            # Found a MILESTONE row - scan its columns to find where dates start
            for col_idx in range(week_col_start, len(row)):
                cell = row[col_idx]
                if isinstance(cell, (datetime.datetime, datetime.date)):
                    # Found the first date column - this is where weeks start
                    week_col_start = col_idx
                    break
            break  # Use the first MILESTONE row we find

    # ── Step 1: build per-milestone week date map ─────────────────────────────
    # Each MILESTONE row has its own week sequence starting from week_col_start.
    # We walk all rows once, and every time we hit a MILESTONE we record its
    # week dates. TASK rows then use the week dates of their current milestone.
    #
    # milestone_weeks: { milestone_name: ['YYYY-MM-DD', ...] }
    milestone_weeks = {}

    for row in all_rows[1:]:  # Skip header row
        row_type = row[col_mapping["type"]]
        # Normalize row type
        row_type = str(row_type).strip().upper() if row_type else ""
        
        if row_type == "MILESTONE":
            name = row[col_mapping["name"]]
            dates = []
            for v in row[week_col_start:]:
                if isinstance(v, (datetime.datetime, datetime.date)):
                    dates.append(date_key(v))
            if dates:
                milestone_weeks[name] = dates

    if not milestone_weeks:
        raise ValueError(
            "No MILESTONE row with week dates found. "
            f"Check that columns starting from column {col_index_to_letter(week_col_start + 1)} contain dates."
        )

    # ── Step 2: walk every row ────────────────────────────────────────────────
    tickets    = []
    warnings   = []
    epics      = {}        # name → milestone
    milestones = []
    current_milestone_weeks = []   # week dates of the milestone currently in scope

    for row_num, row in enumerate(all_rows[1:], start=2):  # Start from row 2 (skip header)
        row_type = row[col_mapping["type"]]
        
        # Normalize row type to uppercase for case-insensitive matching
        row_type = str(row_type).strip().upper() if row_type else ""

        if row_type == "MILESTONE":
            m_name = row[col_mapping["name"]]
            milestones.append(m_name)
            # Switch active week dates to this milestone's sequence
            current_milestone_weeks = milestone_weeks.get(m_name, [])
            continue

        if row_type == "EPIC":
            epics[row[col_mapping["name"]]] = row[col_mapping["milestone"]]
            continue

        if row_type != "TASK":
            continue  # skip header row, totals row, blank rows

        # ── Extract left-column fields ────────────────────────────────────────
        name        = row[col_mapping["name"]]
        description = row[col_mapping["description"]] if col_mapping["description"] < len(row) else None
        milestone   = row[col_mapping["milestone"]] if col_mapping["milestone"] < len(row) else None
        epic        = row[col_mapping["epic"]] if col_mapping["epic"] < len(row) else None
        # Priority: only extract if column exists, otherwise default to "medium"
        priority    = None
        if "priority" in col_mapping and col_mapping["priority"] < len(row):
            priority = row[col_mapping["priority"]]
        if not priority:
            priority = "Medium"  # Default priority if not specified
        effort_raw  = row[col_mapping["effort_hrs"]] if col_mapping["effort_hrs"] < len(row) else None
        assignee    = row[col_mapping["assignee"]] if col_mapping["assignee"] < len(row) else None

        effort_hrs  = parse_effort(effort_raw)
        assignee    = assignee.strip() if assignee else None

        # Use the week dates belonging to this task's milestone.
        # current_milestone_weeks is already set to the right milestone
        # because we process rows top-to-bottom and update on every MILESTONE row.
        week_dates    = current_milestone_weeks
        week_col_count = len(week_dates)

        # ── Extract week hours (cols from week_col_start onward) ────────────────
        week_hours = {}   # { 'YYYY-MM-DD': hours_float }
        week_vals  = row[week_col_start:]   # 0-based

        planned_total = 0.0
        for i, val in enumerate(week_vals[:week_col_count]):
            if val is not None:
                try:
                    hrs = float(val)
                    if hrs > 0:
                        week_hours[week_dates[i]] = hrs
                        planned_total += hrs
                except (ValueError, TypeError):
                    pass

        # ── Warnings ──────────────────────────────────────────────────────────
        if not assignee:
            warnings.append({
                "row":     row_num,
                "task":    name,
                "issue":   "unassigned",
                "detail":  "Assignee is blank. Task cannot be scheduled or conflicted."
            })

        if effort_hrs is not None and planned_total > 0:
            diff = abs(planned_total - effort_hrs)
            if diff > 0.5:   # tolerance: 0.5 hrs
                warnings.append({
                    "row":    row_num,
                    "task":   name,
                    "issue":  "effort_mismatch",
                    "detail": (
                        f"Effort (hrs) col says {effort_hrs}h "
                        f"but week columns sum to {planned_total}h "
                        f"(diff: {diff}h)."
                    )
                })

        if effort_hrs is not None and planned_total == 0:
            warnings.append({
                "row":    row_num,
                "task":   name,
                "issue":  "no_weeks_planned",
                "detail": f"Effort is {effort_hrs}h but no week columns filled in."
            })

        ticket = {
            "row":          row_num,
            "name":         name,
            "description":  description,
            "milestone":    milestone,
            "epic":         epic,
            "priority":     priority,
            "effort_hrs":   effort_hrs,
            "assignee":     assignee,
            "week_hours":   week_hours,        # { week_date: hrs }
            "planned_total": planned_total,    # sum of week cols
            "active_weeks": sorted(week_hours.keys()),
        }
        
        # Debug: log first few tickets to verify priority extraction
        if row_num <= 5:
            logger.debug(f"ROW {row_num} '{name}':")
            logger.debug(f"  Row type (col {col_mapping.get('type', '?')}): {row[col_mapping['type']] if col_mapping['type'] < len(row) else 'N/A'}")
            if "priority" in col_mapping:
                priority_col_idx = col_mapping['priority']
                priority_col_letter = col_index_to_letter(priority_col_idx + 1)
                logger.debug(f"  Priority (col {priority_col_letter} idx {priority_col_idx}): '{row[priority_col_idx] if priority_col_idx < len(row) else 'N/A'}' → Extracted as '{priority}'")
            else:
                logger.debug(f"  Priority: NOT FOUND IN COLUMNS → Using default 'medium'")
        
        tickets.append(ticket)

    # ── Step 3: per-assignee schedule ─────────────────────────────────────────
    # schedule[assignee][week_date] = { total_hrs, tasks: [name, ...] }
    schedule = {}

    for t in tickets:
        if not t["assignee"]:
            continue
        dev = t["assignee"]
        if dev not in schedule:
            schedule[dev] = {}
        for wk, hrs in t["week_hours"].items():
            if wk not in schedule[dev]:
                schedule[dev][wk] = {"total_hrs": 0.0, "tasks": []}
            schedule[dev][wk]["total_hrs"] += hrs
            schedule[dev][wk]["tasks"].append(t["name"])

    # Sort each dev's weeks
    schedule = {
        dev: dict(sorted(weeks.items()))
        for dev, weeks in sorted(schedule.items())
    }

    # ── Step 4: conflicts — same assignee, same week, total hrs > threshold ───
    conflicts = []
    for dev, weeks in schedule.items():
        for wk, info in weeks.items():
            if len(info["tasks"]) > 1:
                conflicts.append({
                    "assignee":   dev,
                    "week":       wk,
                    "total_hrs":  info["total_hrs"],
                    "tasks":      info["tasks"],
                    "overbooked": info["total_hrs"] > HOURS_PER_WEEK,
                })

    # ── Step 5: parallel tasks — different assignees, same week ──────────────
    # Build week → list of (task_name, assignee) for tasks with hours that week
    week_task_map = {}
    for t in tickets:
        if not t["assignee"]:
            continue
        for wk in t["active_weeks"]:
            if wk not in week_task_map:
                week_task_map[wk] = []
            week_task_map[wk].append((t["name"], t["assignee"]))

    parallel_tasks = []
    for wk in sorted(week_task_map.keys()):
        entries = week_task_map[wk]
        if len(entries) < 2:
            continue
        # Find pairs with different assignees
        seen = set()
        for i in range(len(entries)):
            for j in range(i + 1, len(entries)):
                n1, a1 = entries[i]
                n2, a2 = entries[j]
                if a1 != a2:
                    pair_key = tuple(sorted([n1, n2]))
                    if pair_key not in seen:
                        seen.add(pair_key)
                        parallel_tasks.append({
                            "week":       wk,
                            "task_a":     n1,
                            "assignee_a": a1,
                            "task_b":     n2,
                            "assignee_b": a2,
                        })

    # ── Step 6: availability — first free week after last task ────────────────
    availability = {}
    for dev, weeks in schedule.items():
        busy_weeks = sorted(weeks.keys())
        if not busy_weeks:
            availability[dev] = {"last_busy_week": None, "first_free_week": None}
            continue

        last_busy = busy_weeks[-1]
        # Next week after last busy
        last_dt   = datetime.date.fromisoformat(last_busy)
        first_free = (last_dt + datetime.timedelta(weeks=1)).isoformat()

        availability[dev] = {
            "last_busy_week": last_busy,
            "first_free_week": first_free,
            "total_tasks": len([t for t in tickets if t["assignee"] == dev]),
            "total_hrs_planned": sum(
                info["total_hrs"] for info in weeks.values()
            ),
        }

    # ── Step 7: assemble output ───────────────────────────────────────────────
    return {
        "meta": {
            "file":          filepath,
            "parsed_at":     datetime.datetime.now().isoformat(),
            "week_range":    {"start": week_dates[0], "end": week_dates[-1]},
            "total_weeks":   len(week_dates),
            "total_tasks":   len(tickets),
            "total_assignees": len(schedule),
        },
        "tickets":        tickets,
        "schedule":       schedule,
        "conflicts":      conflicts,
        "parallel_tasks": parallel_tasks,
        "availability":   availability,
        "warnings":       warnings,
    }


# ── CLI entry point ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Parse AAI Ops roadmap Excel file.")
    parser.add_argument(
        "--file", "-f",
        default="roadmap_template_AAI_Ops.xlsx",
        help="Path to the roadmap .xlsx file"
    )
    parser.add_argument(
        "--section", "-s",
        choices=["meta","tickets","schedule","conflicts","parallel_tasks","availability","warnings"],
        default=None,
        help="Print only one section of the output"
    )
    args = parser.parse_args()

    result = parse(args.file)

    if args.section:
        print(json.dumps(result[args.section], indent=2, default=str))
    else:
        print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()