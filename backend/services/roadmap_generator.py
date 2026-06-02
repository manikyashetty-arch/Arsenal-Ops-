"""
Roadmap Generator Service - Produce a roadmap .xlsx template from a PRD analysis.

The generated file matches the column layout that ``backend/parser.py``
expects on upload, so a user can download the template, edit it, and
re-upload via ``POST /api/roadmap/parse-file`` without any parser changes.

Layout (sheet name contains ``Roadmap``):
    Type | Name | Description | Milestone | Epic | Priority | Effort (hrs) | Assignee | <Mon1> | <Mon2> | ...

Rows:
    - MILESTONE rows carry Monday dates in the week columns they span.
    - EPIC rows fill type/name/milestone only.
    - TASK rows fill the standard fields plus per-week hours; assignee is
      left blank for the user to fill in before re-upload.
"""

import asyncio
import datetime
import io
import json
import os
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from logging_config import setup_logger

logger = setup_logger("roadmap_generator")

HEADER_FIELDS = [
    "Type",
    "Name",
    "Description",
    "Milestone",
    "Epic",
    "Priority",
    "Effort (hrs)",
    "Assignee",
]

# Cap PRD content sent to the LLM, mirroring architecture_generator.analyze_prd.
PRD_CONTENT_LIMIT = 8000

_client = None


def get_openai_client():
    global _client
    if _client is None:
        try:
            from openai import OpenAI

            _client = OpenAI(
                api_key=os.getenv("OPENAI_API_KEY", ""),
                timeout=90.0,
            )
        except Exception as e:
            logger.warning(f"Failed to initialize OpenAI client: {e}")
            _client = None
    return _client


MODEL_NAME = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

PRIORITY_VALUES = {"high": "High", "medium": "Medium", "low": "Low"}


def _snap_to_monday(d: datetime.date) -> datetime.date:
    """Return the Monday of the week containing ``d`` (weekday 0 == Monday)."""
    return d - datetime.timedelta(days=d.weekday())


def build_week_dates(start_date: datetime.date, end_date: datetime.date) -> list[datetime.date]:
    """Mondays from the week of start_date through the week of end_date, inclusive."""
    if end_date < start_date:
        raise ValueError("end_date must be on or after start_date")
    start = _snap_to_monday(start_date)
    end = _snap_to_monday(end_date)
    weeks = []
    cur = start
    while cur <= end:
        weeks.append(cur)
        cur += datetime.timedelta(days=7)
    return weeks


class RoadmapGenerator:
    """Generate roadmap suggestions from a PRD analysis and render to .xlsx."""

    def __init__(self):
        self.model = MODEL_NAME

    @property
    def client(self):
        return get_openai_client()

    async def generate_suggestions(
        self,
        prd_analysis: dict[str, Any],
        project_name: str,
        week_dates: list[datetime.date],
        sprint_weeks: int,
    ) -> dict[str, Any]:
        """
        Ask the LLM for a structured roadmap proposal based on the PRD analysis.

        Returns a dict with milestones / epics / tasks. Week assignments are
        constrained to ``week_dates`` so the renderer can place values in the
        correct columns deterministically.
        """
        if not self.client:
            raise ValueError("LLM client not initialized. Check OPENAI_API_KEY.")

        week_iso = [w.isoformat() for w in week_dates]
        total_weeks = len(week_iso)

        prd_content = prd_analysis.get("prd_content") or ""
        summary = prd_analysis.get("summary") or ""
        key_features = prd_analysis.get("key_features") or []
        technical_requirements = prd_analysis.get("technical_requirements") or []
        timeline = prd_analysis.get("timeline") or []

        prompt = f"""You are an expert technical project manager. Build a starter roadmap
that an engineering team can use as a planning template.

PROJECT: {project_name}

PRD SUMMARY:
{summary}

KEY FEATURES:
{json.dumps(key_features, indent=2)}

TECHNICAL REQUIREMENTS:
{json.dumps(technical_requirements, indent=2)}

PRD-DERIVED PHASES (informational, may be loose):
{json.dumps(timeline, indent=2)}

PRD CONTENT (truncated):
{prd_content[:PRD_CONTENT_LIMIT]}

PLANNING WINDOW:
- Total weeks available: {total_weeks}
- Sprint length: {sprint_weeks} weeks
- Week start dates (Mondays, ISO format): {json.dumps(week_iso)}

INSTRUCTIONS:
1. Propose 2-5 MILESTONES that partition the planning window. Each milestone
   must reference contiguous weeks from the provided week list — give the
   start_week and end_week as exact ISO dates from that list.
2. Propose 3-10 EPICS total, each assigned to one milestone by name.
3. Propose 8-25 TASKS total, each linked to exactly one epic and one
   milestone. Tasks should be small enough to fit inside their milestone.
4. For each task, give:
   - effort_hrs: integer total hours
   - priority: one of "High", "Medium", "Low"
   - week_hours: object mapping week dates (must be drawn from the provided
     week list AND fall inside the task's milestone) to hours. The sum MUST
     equal effort_hrs. Spread tasks across multiple weeks when realistic.
5. Do NOT assign tasks to people — leave assignees out entirely; the user
   will fill them in before re-uploading.
6. Keep names short (under 80 chars). Descriptions one or two sentences.

Return ONLY this JSON object — no prose, no markdown fences:
{{
  "milestones": [
    {{"name": "STRING", "start_week": "YYYY-MM-DD", "end_week": "YYYY-MM-DD"}}
  ],
  "epics": [
    {{"name": "STRING", "milestone": "STRING", "description": "STRING"}}
  ],
  "tasks": [
    {{
      "name": "STRING",
      "description": "STRING",
      "milestone": "STRING",
      "epic": "STRING",
      "priority": "High|Medium|Low",
      "effort_hrs": NUMBER,
      "week_hours": {{"YYYY-MM-DD": NUMBER, ...}}
    }}
  ]
}}
"""

        client = self.client
        model = self.model

        response = await asyncio.to_thread(
            lambda: client.chat.completions.create(
                model=model,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a project planning assistant. Return only valid JSON "
                            "matching the requested schema."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
                response_format={"type": "json_object"},
                temperature=0.2,
            )
        )

        try:
            parsed = json.loads(response.choices[0].message.content)
        except json.JSONDecodeError as e:
            raise ValueError(f"LLM returned invalid JSON: {e}") from e

        return self._sanitize_suggestions(parsed, week_iso)

    def _sanitize_suggestions(self, raw: dict[str, Any], week_iso: list[str]) -> dict[str, Any]:
        """Defensively normalize LLM output so the renderer never sees junk."""
        week_set = set(week_iso)

        milestones_raw = raw.get("milestones") or []
        epics_raw = raw.get("epics") or []
        tasks_raw = raw.get("tasks") or []

        milestones: list[dict[str, Any]] = []
        for m in milestones_raw:
            if not isinstance(m, dict):
                continue
            name = (m.get("name") or "").strip()
            sw = m.get("start_week")
            ew = m.get("end_week")
            if not name or sw not in week_set or ew not in week_set:
                continue
            if sw > ew:
                sw, ew = ew, sw
            milestones.append({"name": name, "start_week": sw, "end_week": ew})

        if not milestones:
            milestones.append(
                {
                    "name": "Phase 1",
                    "start_week": week_iso[0],
                    "end_week": week_iso[-1],
                }
            )

        milestone_names = {m["name"] for m in milestones}

        # Pre-compute the week list belonging to each milestone for quick lookups.
        idx_by_week = {w: i for i, w in enumerate(week_iso)}
        milestone_week_ranges: dict[str, set[str]] = {}
        for m in milestones:
            i_start = idx_by_week[m["start_week"]]
            i_end = idx_by_week[m["end_week"]]
            milestone_week_ranges[m["name"]] = set(week_iso[i_start : i_end + 1])

        epics: list[dict[str, Any]] = []
        for e in epics_raw:
            if not isinstance(e, dict):
                continue
            name = (e.get("name") or "").strip()
            milestone = (e.get("milestone") or "").strip()
            if not name:
                continue
            if milestone not in milestone_names:
                milestone = milestones[0]["name"]
            epics.append(
                {
                    "name": name,
                    "milestone": milestone,
                    "description": (e.get("description") or "").strip(),
                }
            )

        epic_names = {ep["name"] for ep in epics}
        epic_milestone_lookup = {ep["name"]: ep["milestone"] for ep in epics}

        tasks: list[dict[str, Any]] = []
        for t in tasks_raw:
            if not isinstance(t, dict):
                continue
            name = (t.get("name") or "").strip()
            if not name:
                continue

            epic = (t.get("epic") or "").strip()
            if epic not in epic_names:
                # Drop tasks that don't belong to a known epic — clearer than
                # silently re-parenting them to an arbitrary epic.
                continue

            milestone = epic_milestone_lookup[epic]
            allowed_weeks = milestone_week_ranges[milestone]

            priority_raw = (t.get("priority") or "").strip().lower()
            priority = PRIORITY_VALUES.get(priority_raw, "Medium")

            week_hours_raw = t.get("week_hours") or {}
            week_hours: dict[str, float] = {}
            if isinstance(week_hours_raw, dict):
                for wk, hrs in week_hours_raw.items():
                    if wk not in allowed_weeks:
                        continue
                    try:
                        hrs_f = float(hrs)
                    except (TypeError, ValueError):
                        continue
                    if hrs_f > 0:
                        week_hours[wk] = hrs_f

            # Reconcile effort_hrs with the sum of week_hours so the parser's
            # effort_mismatch warning never fires on re-upload.
            planned_total = sum(week_hours.values())
            try:
                effort_raw = t.get("effort_hrs")
                effort_hrs = float(effort_raw) if effort_raw is not None else 0.0
            except (TypeError, ValueError):
                effort_hrs = 0.0

            if week_hours:
                effort_hrs = planned_total
            elif effort_hrs > 0:
                # No week breakdown given — stuff all hours into the milestone's
                # first week so the file is still re-parseable.
                first_week = sorted(allowed_weeks)[0]
                week_hours[first_week] = effort_hrs

            tasks.append(
                {
                    "name": name,
                    "description": (t.get("description") or "").strip(),
                    "milestone": milestone,
                    "epic": epic,
                    "priority": priority,
                    "effort_hrs": effort_hrs,
                    "week_hours": week_hours,
                }
            )

        return {"milestones": milestones, "epics": epics, "tasks": tasks}

    def build_xlsx(
        self,
        suggestions: dict[str, Any],
        week_dates: list[datetime.date],
    ) -> bytes:
        """Render suggestions to a workbook that parser.parse() can read back.

        Layout note: ``backend/parser.py`` reads each task row's week cells
        relative to its milestone's date count — column 0 of the week area
        means "this milestone's first week", regardless of where in the
        global timeline that milestone sits. So the renderer writes each
        milestone section's rows (MILESTONE + EPICs + TASKs) with values
        starting at the LEFTMOST week column of the row, using local
        per-milestone week indices.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roadmap"

        week_iso = [w.isoformat() for w in week_dates]
        idx_by_week = {w: i for i, w in enumerate(week_iso)}

        # Resolve each milestone's contiguous week list once.
        milestone_week_lookup: dict[str, list[str]] = {}
        for m in suggestions["milestones"]:
            i_start = idx_by_week[m["start_week"]]
            i_end = idx_by_week[m["end_week"]]
            milestone_week_lookup[m["name"]] = week_iso[i_start : i_end + 1]

        # Index epics and tasks by milestone so we can emit one section per
        # milestone in declaration order.
        epics_by_milestone: dict[str, list[dict[str, Any]]] = {}
        for ep in suggestions["epics"]:
            epics_by_milestone.setdefault(ep["milestone"], []).append(ep)

        tasks_by_milestone: dict[str, list[dict[str, Any]]] = {}
        for t in suggestions["tasks"]:
            tasks_by_milestone.setdefault(t["milestone"], []).append(t)

        max_milestone_weeks = max(
            (len(weeks) for weeks in milestone_week_lookup.values()), default=1
        )

        # ── Header row ────────────────────────────────────────────────────
        # Week-column header labels are advisory; parser reads dates from
        # MILESTONE rows, not the header. Use "Week N" so a milestone whose
        # actual weeks differ from the global timeline doesn't look wrong.
        header = HEADER_FIELDS + [f"Week {i + 1}" for i in range(max_milestone_weeks)]
        ws.append(header)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2B2B2B")
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        week_col_start = len(HEADER_FIELDS) + 1  # 1-based

        # ── Per-milestone sections: MILESTONE → EPICs → TASKs ─────────────
        for m in suggestions["milestones"]:
            m_name = m["name"]
            m_weeks = milestone_week_lookup[m_name]

            # MILESTONE row: dates go in the first len(m_weeks) week columns.
            row = ["MILESTONE", m_name, "", "", "", "", "", ""]
            row += [""] * max_milestone_weeks
            ws.append(row)
            ms_row = ws.max_row
            for local_idx, wk_iso in enumerate(m_weeks):
                col = week_col_start + local_idx
                ws.cell(
                    row=ms_row,
                    column=col,
                    value=week_dates[idx_by_week[wk_iso]],
                )
                ws.cell(row=ms_row, column=col).number_format = "yyyy-mm-dd"

            # Group this milestone's tasks by epic so we can write each epic
            # row immediately followed by its own task rows (rather than all
            # epics in a block, then all tasks in a block).
            tasks_by_epic: dict[str, list[dict[str, Any]]] = {}
            for t in tasks_by_milestone.get(m_name, []):
                tasks_by_epic.setdefault(t["epic"], []).append(t)

            def write_task_row(
                t: dict[str, Any],
                *,
                milestone_name: str = m_name,
                milestone_weeks: list[str] = m_weeks,
            ) -> None:
                row = [
                    "TASK",
                    t["name"],
                    t.get("description", ""),
                    milestone_name,
                    t["epic"],
                    t["priority"],
                    t["effort_hrs"],
                    # AI-seeded suggestions deliberately leave assignee empty;
                    # the no-PRD scaffold passes a placeholder name through so
                    # the user can see the column shape.
                    t.get("assignee", ""),
                ]
                # Map global week_hours to this milestone's local indices.
                week_cells: list[Any] = [""] * max_milestone_weeks
                for local_idx, wk_iso in enumerate(milestone_weeks):
                    hrs = t["week_hours"].get(wk_iso)
                    if hrs:
                        week_cells[local_idx] = hrs
                row += week_cells
                ws.append(row)

            # For each EPIC under this milestone: write the epic row, then
            # immediately write the tasks that belong to it.
            written_epic_names: set[str] = set()
            for ep in epics_by_milestone.get(m_name, []):
                row = [
                    "EPIC",
                    ep["name"],
                    ep.get("description", ""),
                    m_name,
                    "",
                    "",
                    "",
                    "",
                ]
                row += [""] * max_milestone_weeks
                ws.append(row)
                written_epic_names.add(ep["name"])
                for t in tasks_by_epic.get(ep["name"], []):
                    write_task_row(t)

            # Orphan tasks (their epic name didn't appear in this milestone's
            # epic list) — append at the end so they aren't silently dropped.
            # Shouldn't normally happen because _sanitize_suggestions ties tasks
            # to known epics, but be defensive.
            for epic_name, tasks in tasks_by_epic.items():
                if epic_name in written_epic_names:
                    continue
                for t in tasks:
                    write_task_row(t)

        # ── Sizing & freeze ───────────────────────────────────────────────
        col_widths = [12, 36, 48, 22, 26, 10, 12, 22] + [12] * max_milestone_weeks
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = ws.cell(row=2, column=week_col_start)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


roadmap_generator = RoadmapGenerator()
